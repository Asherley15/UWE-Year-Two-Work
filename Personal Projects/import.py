from openpyxl import Workbook, load_workbook
from openpyxl.reader.excel import load_workbook


loc = ("/Users/ashleypearson/Documents/UWE/Year Two/Personal Projects/mileage.xlsx")

month = str(input("What month would you like to check?"))
wb = load_workbook(loc, data_only=True)
sheet = wb[month]
total = (sheet['G57'].value)
print(total)
